def isLetter(char):
	if(char == ',' or char == ' ' or char == '.'):
		return False
	else:
		return True

def checkTitle(str_input):
	if "Mr." not in str_input and "Ms." not in str_input and "Mrs." not in str_input:
		return False
	else:
		return True

def checkSuffix(str_input):
	str_in = str_input
	str_lwr = str_in.lower()
	index = -1
	counter = 1
	new_str = ""
	found = False

	for x in range(len(str_in) - 2):
		if (str_lwr[x] == 'i' and str_lwr[x + 1] == 'n' and str_lwr[x + 2] == 'c' and (isLetter(str_lwr[x - 1]) == False)):
			index = x
			found = True
			break
		elif (str_lwr[x] == 'l' and str_lwr[x + 1] == 'l' and str_lwr[x + 2] == 'c' and (isLetter(str_lwr[x - 1]) == False)):
			index = x
			found = True
			break
		elif (str_lwr[x] == 'l' and str_lwr[x + 1] == 't' and str_lwr[x + 2] == 'd' and (isLetter(str_lwr[x - 1]) == False)):
			index = x
			found = True
			break
		elif (str_lwr[x] == 'l' and str_lwr[x + 1] == '.' and str_lwr[x + 2] == 'l' and str_lwr[x + 3] == '.'
				and str_lwr[x + 4] == 'c' and (isLetter(str_lwr[x - 1]) == False)):
			index = x
			found = True
			break

	if found == True:	
		while ((index - counter) >= 0) and (isLetter(str_in[index - counter]) == False):
			counter = counter + 1
			
		new_str = str_in[0:(index - (counter - 1))]
	else:
		new_str = str_in

	return new_str

#Openpyxl library is used for excel sheet manipulation
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
time = datetime.datetime.now()

#Loading current export excel sheet
print(" -------------------------")
print(" Welcome to Filtren v1.3.2 ")
print(" -------------------------\n")

found_file = False

while found_file == False:
	file_in = input("Please enter file name: ")

	if ".xlsx" not in file_in:
		file_in = file_in + ".xlsx"

	try:
	    source = load_workbook(file_in)
	    found_file = True
	except FileNotFoundError as not_found:
		print('WARNING! \"' + not_found.filename + '\" not found, please try again.\n')

sheet = source.active

#Creating new Workbook for filtered info to be saved
new_wb = Workbook(write_only=True)
wb_sheet = new_wb.create_sheet()

#Printing message to user
filter_choice = -1
valid_choice = False

while valid_choice == False:
	filter_choice = input("\nDepartment filtering options:\n   1 - Sales\n   2 - HR/Finance\n   3 - Management\n   4 - Sales/Management (including blank departments)\n   5 - No department filter (including blank departments)\n   6 - No department filter (NOT including blank departments)\n\nPlease type a number: ")
	if filter_choice == '1' or filter_choice == '2' or filter_choice == '3' or filter_choice == '4' or filter_choice == '5' or filter_choice == '6':
		valid_choice = True
	else:
		print("Error! Invalid option, please try again.\n")

if filter_choice == '1':
	print("\n[Sales] Filtering process started ...")
elif filter_choice == '2':
	print("\n[HR/Finance] Filtering process started ...")
elif filter_choice == '3':
	print("\n[Management] Filtering process started ...")
elif filter_choice == '4':
	print("\n[Sales/Management/Blanks] Filtering process started ...")
elif filter_choice == '5':
	print("\n[No Department/Blanks] Filtering process started ...")
else:
	print("\n[No Department/No Blanks] Filtering process started ...")

#Initializing variables to keep track of rows and criteria
currentRow = 1
filteredSheetRow = 1
is_empty = False
status_warning = False
loop_stopped = False

#Indexes for header info
name_index = -1
email_index = -1
first_name_index = -1
last_name_index = -1
department_index = -1
title_index = -1
status_index = -1

#Putting header at top of file
header = ("Company Name", "Title", "First Name", "Last Name", "Email")
wb_sheet.append(header)

found_header = False

for eachRow in sheet.iter_rows():
	#Resetting variables
	is_empty = False

	if found_header == False:
		for x in range(1, 50):
			headerValue = sheet.cell(row=currentRow, column=x).value
			if headerValue != None:
				headerValue = headerValue.lower()
				if headerValue == "rname" and name_index == -1:
					name_index = x
				elif headerValue == "cpemail" and email_index == -1:
					email_index = x
				elif headerValue == "cpfirstname" and first_name_index == -1:
					first_name_index = x
				elif headerValue == "cpsurname" and last_name_index == -1:
					last_name_index = x
				elif headerValue == "cpdepartment" and department_index == -1:
					department_index = x
				elif headerValue == "status" and status_index == -1:
					status_index = x
				elif headerValue == "cptitle" and title_index == -1:
					title_index = x

			# 	print("X: " + str(x) + ", [" + headerValue + "]")
			# else:
			# 	print("X: " + str(x) + ", [NONE]")

		found_header = True

	if name_index == -1:
		print("\nERROR! Missing \"RName\" column in file.\n")
		loop_stopped = True
		break
	elif email_index == -1:
		print("\nERROR! Missing \"Email\" column in file.\n")
		loop_stopped = True
		break
	elif first_name_index == -1:
		print("\nERROR! Missing \"CPFirstName\" column in file.\n")
		loop_stopped = True
		break
	elif last_name_index == -1:
		print("\nERROR! Missing \"CPSurName\" column in file.\n")
		loop_stopped = True
		break
	elif department_index == -1:
		print("\nERROR! Missing \"CPDepartment\" column in file.\n")
		loop_stopped = True
		break
	elif status_index == -1:
		print("\nERROR! Missing \"Status\" column in file.\n")
		loop_stopped = True
		break
	elif title_index == -1:
		print("\nERROR! Missing \"Title\" column in file.\n")
		loop_stopped = True
		break

	# Variables for reference
	# name_index
	# email_index
	# first_name_index
	# last_name_index
	# department_index
	# status_index

	#Parsing relevant information from export
	company_name = sheet.cell(row=currentRow, column=name_index).value
	email = sheet.cell(row=currentRow, column=email_index).value
	title = sheet.cell(row=currentRow, column=title_index).value
	first_name = sheet.cell(row=currentRow, column=first_name_index).value
	last_name = sheet.cell(row=currentRow, column=last_name_index).value
	department = sheet.cell(row=currentRow, column=department_index).value
	status = sheet.cell(row=currentRow, column=status_index).value

	#Checking if fields are blank, if so then flag will be set
	if company_name is None:
		is_empty = True
	elif email is None:
		is_empty = True
	elif title is None:
		is_empty = True
	elif first_name == None:
		is_empty = True
	elif last_name == None:
		is_empty = True
	elif department == None:
		if filter_choice != '4' and filter_choice != '5':
			is_empty = True
	elif status == None:
		is_empty = True

	#Only for testing purposes
	#print("[" + str(company_name) + ", " + str(email) + ", " + str(status) + ", Empty Status: " + str(is_empty) + "]")
	#print(str(email))

	#Generating department warning
	# if (department != "Sales" and department_warning == False):
	# 	#print('\x1b[1;37;41m' + 'WARNING! This export contains other departments besides Sales, do you wish to continue?' + '\x1b[0m')
	# 	input_value = input("Type \'y\' for YES or \'n\' for NO: ")
	# 	department_warning = True
	# 	#Handling department warning response
	# 	if input_value == 'n':
	# 		print('\x1b[0;30;43m' + 'Filtering process aborted.' + '\x1b[0m')
	# 		loop_stopped = True
	# 		break	

	#Checking details in variables
	if is_empty == False:
		#Generating status warning
		if status_warning == False and status.lower() != "mailing list" and status.lower() != "status":
			print('\nMultiple entries without type \'Mailing List\' detected and will be filtered out, do you wish to continue?')
			status_warning_response = input("Type \'y\' for YES or \'n\' for NO: ")

			#Handling status warning response
			if status_warning_response == 'n':
				print('\nFiltering process terminated.\n')
				loop_stopped = True
				break
			else:
				print("\n")
				status_warning = True

		if checkTitle(title) == True:
			#Creating tuple to be written to file if valid
			temp_tup = (checkSuffix(company_name), title, first_name, last_name, email)

			#Making sure email address is valid
			if '@' in email:
				if status.lower() == "mailing list":
					if filter_choice == '1':
						if department.lower() == "sales":
							wb_sheet.append(temp_tup)
					elif filter_choice == '2':
						if department.lower() == "finance" or department.lower() == "hr":
							wb_sheet.append(temp_tup)
					elif filter_choice == '3':
						if department.lower() == "management":
							wb_sheet.append(temp_tup)
					elif filter_choice == '4':
						if department == None or "management" in department.lower() or "sales" in department.lower():
							wb_sheet.append(temp_tup)
					elif filter_choice == '5' or filter_choice == '6':
							wb_sheet.append(temp_tup)
					
	#Increading row
	currentRow += 1

if loop_stopped == False:
	#Creating the filename

	# department_file = ""

	# if filter_choice == '1':
	# 	department_file = "Sales"
	# else:
	# 	department_file = "Finance/HR"

	#filtered_result_file_name = department_file + " Filtered Result " + str(time.day) + "-" + str(time.month) + "-" + str(time.year) + ".xlsx"
	#filtered_result_file_name =  "Filtered Result " + str(time.day) + "-" + str(time.month) + "-" + str(time.year) + ".xlsx"

	#Saving the new Workbook
	#new_wb.save(filtered_result_file_name)

	#Printing message to user
	print("Filtering complete.")

	saved_file = False

	while saved_file == False:
		file_name = input("\nPlease enter name for new file: ")

		if ".xlsx" not in file_name:
			file_name = file_name + ".xlsx"
		
		print("\nFiltered results will be saved as \"" + str(file_name) + "\", continue?")
		continue_choice = input("Type \'y\' for YES or \'n\' for NO: ")

		if continue_choice == 'y':
			try:
			    temp_src = load_workbook(file_name)
			    print("\nWARNING! \"" + str(file_name) + "\" already exists in current directory. Do you want to overwrite it?")
			    overwrite_choice = input("Type \'y\' for YES or \'n\' for NO: ")

			    if overwrite_choice == 'y':
			    	print("\n\"" + str(file_name) + "\" saved successfully!\n")
			    	new_wb.save(file_name)
			    	saved_file = True
			except FileNotFoundError as not_found:
				print("\n\"" + str(file_name) + "\" saved successfully!\n")
				new_wb.save(file_name)
				saved_file = True
	

else:
	err_tup = ('FILTERING WAS INTERRUPTED, DO NOT USE THIS FILE!', 'FILTERING WAS INTERRUPTED, DO NOT USE THIS FILE', 'FILTERING WAS INTERRUPTED, DO NOT USE THIS FILE!', 'FILTERING WAS INTERRUPTED, DO NOT USE THIS FILE!', 'FILTERING WAS INTERRUPTED, DO NOT USE THIS FILE!')
	wb_sheet.append(err_tup)
	new_wb.save("FILTERING ERROR, DO NOT USE.xlsx")

while True:
		input("Press \"enter\" to exit.")
		break


